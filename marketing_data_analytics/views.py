import matplotlib.pyplot as plt

# 日本語フォントを指定してフォント設定
plt.rcParams['font.sans-serif'] = ['Yu Gothic', 'Meiryo', 'Takao', 'IPAexGothic', 'IPAPGothic', 'VL PGothic', 'Noto Sans CJK JP']
plt.rcParams['axes.unicode_minus'] = False  # マイナス記号が文字化けするのを防ぐ

# 以降のコードはそのまま
import matplotlib.font_manager as fm
from django.shortcuts import render
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
import pandas as pd
import os
import io
from django.http import HttpResponse
from django.conf import settings
from .models import Review
import japanize_matplotlib


# 年齢を代表的な年齢区間に変換する関数
def convert_to_age_group(age_series):
    return age_series.apply(lambda age: f'{age // 10 * 10}代')

def create_scatter_plot(x_data, y_data, x_label, y_label):
    # 以前のフォントの設定をクリアしてデフォルトのサンセリフフォントを使用
    plt.rcParams['font.sans-serif'] = []

    # 散布図を作成する関数
    fig, ax = plt.subplots()
    ax.scatter(x_data, y_data)
    ax.set_xlabel(x_label)
    ax.set_ylabel(y_label)
    ax.set_title(f"{x_label} vs {y_label}")

    # 画像を保存
    img_data = io.BytesIO()
    plt.savefig(img_data, format='png')
    plt.close()

    img_data.seek(0)  # ファイルポインタを先頭に戻す

    return img_data

def top_page(request):
    return render(request, 'top_page.html')

def save_data(request):
    if request.method == 'POST':
        form = request.POST
        age = form.get('age')
        gender = form.get('gender')
        menu = form.get('menu')
        overall_satisfaction = form.get('overall_satisfaction')
        food_satisfaction = form.get('food_satisfaction')
        price_satisfaction = form.get('price_satisfaction')
        ambience_satisfaction = form.get('ambience_satisfaction')
        service_satisfaction = form.get('service_satisfaction')
        other_comments = form.get('other_comments')

        # フォームのデータを辞書としてまとめる
        data = {
            'age': age,
            'gender': gender,
            'menu': menu,
            'overall_satisfaction': overall_satisfaction,
            'food_satisfaction': food_satisfaction,
            'price_satisfaction': price_satisfaction,
            'ambience_satisfaction': ambience_satisfaction,
            'service_satisfaction': service_satisfaction,
            'other_comments': other_comments,
        }

        review = Review.objects.create(**data)

        return render(request, 'success_page.html')
    else:
        return HttpResponse("無効なリクエストメソッドです。POSTメソッドを使用してください。")

def success_page(request):
    return render(request, 'success_page.html')

def export_to_excel(request):
    data = Review.objects.all()

    # データをPandas DataFrameに変換
    data_df = pd.DataFrame(list(data.values()))

    # 数値データのみを含むDataFrameを作成
    numerical_data_df = data_df[['age', 'overall_satisfaction', 'food_satisfaction', 'price_satisfaction', 'ambience_satisfaction', 'service_satisfaction']]

    # 平均値を計算
    average_data = numerical_data_df.mean().round(2)

    # 相関行列を計算
    correlation_matrix = numerical_data_df.corr().round(2)

    # デモグラフィックデータを準備
    demographic_data = data_df[['gender', 'age']].copy()
    demographic_data['age'] = convert_to_age_group(demographic_data['age'])
    # print(demographic_data)
    # デモグラフィック分析：各客層ごとの人数と全体の割合を計算
    demographic_analysis = demographic_data.groupby(['gender', 'age']).size()
    total_count = demographic_analysis.sum()
    demographic_analysis = demographic_analysis.reset_index(name='人数')

    demographic_analysis['全体の割合'] = (demographic_analysis['人数'] / total_count * 100).round(2)
    # print(demographic_analysis)

# すべての客層を表現するDataFrameを作成
    genders = ['男性', '女性']
    ages = [f'{i}代' for i in range(10, 90, 10)]
    all_demographics = pd.DataFrame([(gender, age) for gender in genders for age in ages], columns=['gender', 'age'])

# 欠損値を補完する
    missing_rows = []
    for gender in genders:
        for age in ages:
            if not ((demographic_analysis['gender'] == gender) & (demographic_analysis['age'] == age)).any():
                missing_rows.append({'gender': gender, 'age': age, '人数': 0, '全体の割合': 0})

# デモグラフィック分析結果をマージ
    if missing_rows:
        missing_df = pd.DataFrame(missing_rows)
        demographic_analysis = pd.concat([demographic_analysis, missing_df])

# デモグラフィック分析結果をマージ
    result_df = pd.merge(all_demographics, demographic_analysis, how='left', on=['age', 'gender'])
    result_df['人数'] = result_df['人数'].fillna(0)
    result_df['全体の割合'] = result_df['全体の割合'].fillna(0)
    # データを年代でソート
    result_df['age'] = pd.Categorical(result_df['age'], categories=ages, ordered=True)
    result_df.sort_values(['gender', 'age'], inplace=True)
    result_df.reset_index(drop=True, inplace=True)

    # Excelファイルに出力（ファイル名を変更）
    try:
        file_path = os.path.join(settings.MEDIA_ROOT, 'data_analytics.xlsx')
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            average_data.to_excel(writer, sheet_name='Average')
            correlation_matrix.to_excel(writer, sheet_name='CorrelationMatrix')
            result_df.to_excel(writer, sheet_name='DemographicAnalysis', index=False)  # デモグラフィック分析結果を出力

            # 散布図を追加していく処理
            scatter_plots = {
                '料理の満足度': numerical_data_df['food_satisfaction'],
                '価格の満足度': numerical_data_df['price_satisfaction'],
                '店の雰囲気': numerical_data_df['ambience_satisfaction'],
                '接客態度': numerical_data_df['service_satisfaction'],
            }

            for label, data in scatter_plots.items():
                img_data = create_scatter_plot(numerical_data_df['overall_satisfaction'], data, 'Overall Satisfaction', label)
                # 新しいWorksheetを作成してそこに散布図を追加
                scatter_sheet = writer.book.create_sheet(title=f'{label}_vs_Overall_Satisfaction')
                scatter_sheet.add_image(Image(img_data), 'A1')

            # 相関行列のシートを開いてセルの色を設定
            sheet = writer.book['CorrelationMatrix']

            # セルの色を設定するためのカスタム関数
            def set_cell_color(cell, color):
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            '''
            for row in sheet.iter_rows(min_row=2, max_row=correlation_matrix.shape[0]+1, min_col=2, max_col=correlation_matrix.shape[1]+1):
                for cell in row:
                    value = cell.value
                    if value < -0.7:
                        set_cell_color(cell, "6495ED")  # 負の相関で濃い青に設定
                    elif value < -0.5:
                        set_cell_color(cell, "ADD8E6")  # 負の相関でもう少し濃い青に設定
                    elif value < -0.3:
                        set_cell_color(cell, "E0F8FF")  # 負の相関でさらに薄い青に設定
                    elif 0.3 <= value < 0.5:
                        set_cell_color(cell, "FFEBEB")  # 正の相関でオレンジ色に設定
                    elif 0.3 <= value < 0.7:
                        set_cell_color(cell, "FFB6B6")  # 正の相関で金色に設定
                    elif 0.7 <= value < 0.99:
                        set_cell_color(cell, "FF8282")  # 正の相関でダークオレンジ色に設定
                        '''

            week_correlation_cell = sheet.cell(row=correlation_matrix.shape[0]+3, column=2)  
            week_correlation_cell.value = "相関係数-0.3以下-0.5より大きいで弱い負の相関"
            set_cell_color(week_correlation_cell, "E0F8FF")  # 負の相関で薄い青色に設定

            normal_correlation_cell = sheet.cell(row=correlation_matrix.shape[0]+4, column=2)
            normal_correlation_cell.value = "相関係数-0.5以下-0.7より大きいで普通の負の相関"
            set_cell_color(normal_correlation_cell, "ADD8E6")  # 負の相関で青色に設定

            strong_correlation_cell = sheet.cell(row=correlation_matrix.shape[0]+5, column=2)
            strong_correlation_cell.value = "相関係数-0.7以下で強い負の相関"
            set_cell_color(strong_correlation_cell, "6495ED")  # 負の相関で濃い青に設定

            strong_positive_correlation_cell = sheet.cell(row=correlation_matrix.shape[0]+6, column=2)
            strong_positive_correlation_cell.value = "相関係数0.7以上で強い正の相関"
            set_cell_color(strong_positive_correlation_cell, "FF8282")  # 正の相関で濃い赤色に設定

            normal_positive_correlation_cell = sheet.cell(row=correlation_matrix.shape[0]+7, column=2)
            normal_positive_correlation_cell.value = "相関係数0.5以上0.7未満で普通の正の相関"
            set_cell_color(normal_positive_correlation_cell, "FFB6B6")  # 正の相関で赤色に設定

            weak_positive_correlation_cell = sheet.cell(row=correlation_matrix.shape[0]+8, column=2)
            weak_positive_correlation_cell.value = "相関係数0.3以上0.5未満で弱い正の相関"
            set_cell_color(weak_positive_correlation_cell, "FFEBEB")  # 正の相関で薄い赤色に設定
        # ファイルをダウンロードさせるResponseオブジェクトを作成
        with open(file_path, 'rb') as file:
            response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=data_analytics.xlsx'

        return response
    # return HttpResponse("エクセルファイルの出力完了")
    except Exception as e:
        return HttpResponse(e)

